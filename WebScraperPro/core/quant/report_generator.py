"""
Quantitative Finance Report Generator
=====================================
Generates professional PDF and Excel reports from QuantEngine analysis results.

Provides two main classes:
    - PDFReportGenerator: Produces formatted PDF documents with cover page,
      table of contents, and per-analysis sections using ReportLab.
    - ExcelReportGenerator: Produces multi-sheet Excel workbooks with summary
      and detailed per-category sheets using openpyxl.

Typical usage::

    from core.quant.report_generator import PDFReportGenerator, ExcelReportGenerator

    pdf_gen = PDFReportGenerator()
    pdf_path = pdf_gen.generate_report(engine.history, "analysis.pdf")

    xls_gen = ExcelReportGenerator()
    xls_path = xls_gen.generate_workbook(engine.history, "analysis.xlsx")
"""

from __future__ import annotations

import os
import math
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import numpy as np

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    KeepTogether, HRFlowable,
)
from reportlab.pdfgen import canvas as pdfcanvas

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _flatten_dict(
    obj: Any,
    parent_key: str = "",
    sep: str = ".",
) -> Dict[str, Any]:
    """Recursively flatten a nested dictionary into dot-notation keys.

    Examples
    --------
    >>> _flatten_dict({"call": {"price": 12.5, "delta": 0.6}})
    {'call.price': 12.5, 'call.delta': 0.6}
    >>> _flatten_dict([1, 2, 3])
    {'value': '[1, 2, 3]'}
    """
    items: List[Tuple[str, Any]] = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else str(k)
            items.extend(_flatten_dict(v, new_key, sep).items())
    else:
        items.append((parent_key or "value", _serialise(obj)))
    return dict(items)


def _serialise(val: Any) -> Any:
    """Convert a value to a JSON-safe, human-readable representation.

    NumPy arrays are converted to nested Python lists.  Other
    non-serialisable types fall back to ``str()``.
    """
    if isinstance(val, np.ndarray):
        return val.tolist()
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        return float(val)
    if isinstance(val, (np.bool_,)):
        return bool(val)
    if isinstance(val, dict):
        return {str(k): _serialise(v) for k, v in val.items()}
    if isinstance(val, (list, tuple)):
        return [_serialise(v) for v in val]
    try:
        # Quick check: does it survive round-trip?
        import json
        json.dumps(val)
        return val
    except (TypeError, ValueError, OverflowError):
        return str(val)


def _fmt_val(val: Any, max_len: int = 60) -> str:
    """Format a value for display in table cells.

    Numbers are rounded to 4 significant figures when possible.
    Long strings are truncated with an ellipsis.
    """
    if isinstance(val, float):
        return f"{val:.4g}"
    if isinstance(val, (int, np.integer)):
        return str(int(val))
    s = str(val)
    if len(s) > max_len:
        return s[: max_len - 3] + "..."
    return s


def _excel_val(val: Any) -> Any:
    """Convert a value for safe insertion into an Excel cell.

    Lists, dicts and other non-scalar types are stringified.
    NumPy scalars are unwrapped to plain Python types.
    """
    if val is None:
        return None
    if isinstance(val, (list, tuple, dict, set)):
        return _fmt_val(val, max_len=200)
    if isinstance(val, np.ndarray):
        return _fmt_val(val, max_len=200)
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        return float(val)
    if isinstance(val, (np.bool_,)):
        return bool(val)
    try:
        import json
        json.dumps(val)
        return val
    except (TypeError, ValueError, OverflowError):
        return str(val)


def _ensure_parent_dir(path: str) -> None:
    """Create parent directories for *path* if they do not exist."""
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)


def _now_str() -> str:
    """Return the current UTC datetime as an ISO-format string."""
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")


def _group_by_category(records: List[Dict]) -> Dict[str, List[Dict]]:
    """Partition *records* into an ordered dict keyed by ``category``."""
    groups: Dict[str, List[Dict]] = {}
    for r in records:
        cat = r.get("category", "Uncategorised")
        groups.setdefault(cat, []).append(r)
    return groups


# ---------------------------------------------------------------------------
# PDF colour / style constants
# ---------------------------------------------------------------------------

_BRAND_BLUE = colors.HexColor("#1a3c6e")
_BRAND_LIGHT = colors.HexColor("#eaf0f9")
_ACCENT = colors.HexColor("#2c7bb6")
_TEXT_DARK = colors.HexColor("#1e1e1e")
_GREY_LINE = colors.HexColor("#cccccc")


# ---------------------------------------------------------------------------
# PDFReportGenerator
# ---------------------------------------------------------------------------

class PDFReportGenerator:
    """Generate professional PDF reports from QuantEngine analysis history.

    Each record in the history is expected to be a dict with keys
    ``category``, ``method``, ``result``, and ``timestamp`` (as produced
    by :meth:`QuantEngine._record`).

    Parameters
    ----------
    page_size : tuple
        ReportLab page size tuple (default A4).
    """

    def __init__(self, page_size: tuple = A4) -> None:
        self._page_size = page_size
        self._styles = getSampleStyleSheet()
        self._styles.add(ParagraphStyle(
            name="CoverTitle",
            fontName="Helvetica-Bold",
            fontSize=28,
            leading=34,
            textColor=_BRAND_BLUE,
            alignment=TA_CENTER,
            spaceAfter=12,
        ))
        self._styles.add(ParagraphStyle(
            name="CoverSubtitle",
            fontName="Helvetica",
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#555555"),
            alignment=TA_CENTER,
            spaceAfter=6,
        ))
        self._styles.add(ParagraphStyle(
            name="SectionHeading",
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            textColor=_BRAND_BLUE,
            spaceBefore=18,
            spaceAfter=8,
        ))
        self._styles.add(ParagraphStyle(
            name="SubHeading",
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=15,
            textColor=_ACCENT,
            spaceBefore=10,
            spaceAfter=4,
        ))
        self._styles.add(ParagraphStyle(
            name="BodyText2",
            fontName="Helvetica",
            fontSize=9,
            leading=13,
            textColor=_TEXT_DARK,
            alignment=TA_JUSTIFY,
        ))
        self._styles.add(ParagraphStyle(
            name="TOCEntry",
            fontName="Helvetica",
            fontSize=11,
            leading=18,
            textColor=_TEXT_DARK,
            leftIndent=20,
        ))
        self._styles.add(ParagraphStyle(
            name="FooterStyle",
            fontName="Helvetica",
            fontSize=8,
            textColor=colors.HexColor("#888888"),
            alignment=TA_CENTER,
        ))

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate_report(
        self,
        results: List[Dict],
        output_path: str,
        title: str = "Quantitative Analysis Report",
    ) -> str:
        """Generate a full PDF report from a list of analysis results.

        Parameters
        ----------
        results : list[dict]
            Each dict must contain ``category``, ``method``, ``result``,
            and ``timestamp`` keys.
        output_path : str
            Filesystem path for the generated PDF.
        title : str
            Report title shown on the cover page.

        Returns
        -------
        str
            The *output_path* on success.

        Raises
        ------
        ValueError
            If *results* is empty.
        RuntimeError
            If PDF generation fails for any reason.
        """
        if not results:
            raise ValueError("No analysis results to report.")
        _ensure_parent_dir(output_path)

        doc = SimpleDocTemplate(
            output_path,
            pagesize=self._page_size,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
            leftMargin=0.9 * inch,
            rightMargin=0.9 * inch,
        )
        story: list = []
        self._build_cover(story, title)
        story.append(PageBreak())
        self._build_toc(story, results)
        story.append(PageBreak())
        self._build_sections(story, results)
        doc.build(story, onFirstPage=self._footer, onLaterPages=self._footer)
        return output_path

    def generate_summary_report(
        self,
        engine_history: List[Dict],
        output_path: str,
    ) -> str:
        """Create a concise summary PDF of all analyses run.

        Produces a single table summarising every analysis entry with
        category, method, timestamp, and the number of metrics.

        Parameters
        ----------
        engine_history : list[dict]
            Full history list from :attr:`QuantEngine.history`.
        output_path : str
            Destination path for the PDF.

        Returns
        -------
        str
            The *output_path* on success.
        """
        if not engine_history:
            raise ValueError("No history entries to summarise.")
        _ensure_parent_dir(output_path)

        doc = SimpleDocTemplate(
            output_path,
            pagesize=self._page_size,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
            leftMargin=0.9 * inch,
            rightMargin=0.9 * inch,
        )
        story: list = []
        story.append(Paragraph("Analysis Summary", self._styles["CoverTitle"]))
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            f"Generated {_now_str()}  |  WebScraper Pro",
            self._styles["CoverSubtitle"],
        ))
        story.append(Spacer(1, 20))

        header = ["#", "Category", "Method", "Timestamp", "Metrics"]
        rows = [header]
        for i, rec in enumerate(engine_history, 1):
            res = rec.get("result", {})
            flat = _flatten_dict(res) if isinstance(res, dict) else {}
            rows.append([
                str(i),
                str(rec.get("category", "")),
                str(rec.get("method", "")),
                str(rec.get("timestamp", ""))[:19],
                str(len(flat)),
            ])

        tbl = Table(rows, colWidths=[30, 110, 140, 120, 50])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _BRAND_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _BRAND_LIGHT]),
            ("GRID", (0, 0), (-1, -1), 0.5, _GREY_LINE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(tbl)
        doc.build(story, onFirstPage=self._footer, onLaterPages=self._footer)
        return output_path

    # ------------------------------------------------------------------
    # Internal: PDF building blocks
    # ------------------------------------------------------------------

    def _build_cover(self, story: list, title: str) -> None:
        """Append cover-page elements to *story*."""
        story.append(Spacer(1, 1.8 * inch))
        story.append(Paragraph("WebScraper Pro", self._styles["CoverSubtitle"]))
        story.append(Spacer(1, 14))
        story.append(HRFlowable(width="60%", thickness=2, color=_BRAND_BLUE,
                               spaceBefore=4, spaceAfter=4))
        story.append(Spacer(1, 14))
        story.append(Paragraph(title, self._styles["CoverTitle"]))
        story.append(Spacer(1, 12))
        story.append(Paragraph(_now_str(), self._styles["CoverSubtitle"]))
        story.append(Spacer(1, 8))
        story.append(Paragraph(
            "Automated Quantitative Finance Report",
            self._styles["CoverSubtitle"],
        ))

    def _build_toc(self, story: list, results: List[Dict]) -> None:
        """Append a table of contents page to *story*."""
        story.append(Paragraph("Table of Contents", self._styles["SectionHeading"]))
        story.append(Spacer(1, 12))
        groups = _group_by_category(results)
        page = 1  # rough counter for TOC display
        for cat, recs in groups.items():
            story.append(Paragraph(f"<b>{cat}</b>", self._styles["TOCEntry"]))
            for rec in recs:
                page += 1
                method = rec.get("method", "")
                story.append(
                    Paragraph(f"\u2003\u2022 {method}", self._styles["TOCEntry"])
                )
            story.append(Spacer(1, 4))

    def _build_sections(self, story: list, results: List[Dict]) -> None:
        """Append one section per category with detail tables."""
        groups = _group_by_category(results)
        for cat, recs in groups.items():
            story.append(Paragraph(cat, self._styles["SectionHeading"]))
            story.append(HRFlowable(width="100%", thickness=1, color=_ACCENT,
                                   spaceBefore=2, spaceAfter=8))
            for rec in recs:
                method = rec.get("method", "Unknown Method")
                ts = rec.get("timestamp", "")
                story.append(
                    Paragraph(f"{method}  <font size=8 color='#888888'>({ts[:19]})</font>",
                              self._styles["SubHeading"])
                )
                res = rec.get("result", {})
                if isinstance(res, dict):
                    flat = _flatten_dict(res)
                elif isinstance(res, (list, tuple)):
                    flat = {"value": _serialise(res)}
                else:
                    flat = {"value": _fmt_val(res)}
                self._add_kv_table(story, flat)
                story.append(Spacer(1, 10))

    @staticmethod
    def _add_kv_table(story: list, flat: Dict[str, Any]) -> None:
        """Append a key-value table for a single analysis result."""
        rows: List[list] = [[Paragraph("<b>Key</b>", ParagraphStyle(
            "kh", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("<b>Value</b>", ParagraphStyle(
            "vh", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white))]]
        for k, v in flat.items():
            rows.append([
                Paragraph(str(k), ParagraphStyle(
                    "kc", fontName="Helvetica", fontSize=8)),
                Paragraph(_fmt_val(v), ParagraphStyle(
                    "vc", fontName="Helvetica", fontSize=8)),
            ])
        tbl = Table(rows, colWidths=[2.2 * inch, 4.0 * inch])
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), _ACCENT),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _BRAND_LIGHT]),
            ("GRID", (0, 0), (-1, -1), 0.4, _GREY_LINE),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]
        tbl.setStyle(TableStyle(style_cmds))
        story.append(tbl)

    @staticmethod
    def _footer(canvas_obj: pdfcanvas.Canvas, doc: SimpleDocTemplate) -> None:
        """Draw page number footer on every page."""
        canvas_obj.saveState()
        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.setFillColor(colors.HexColor("#888888"))
        text = f"WebScraper Pro  |  Page {doc.page}"
        canvas_obj.drawCentredString(A4[0] / 2, 0.45 * inch, text)
        canvas_obj.restoreState()


# ---------------------------------------------------------------------------
# ExcelReportGenerator
# ---------------------------------------------------------------------------

class ExcelReportGenerator:
    """Generate professional Excel workbooks from QuantEngine analysis results.

    The workbook contains a **Summary** sheet with one row per analysis
    plus one additional sheet per *category* holding the detailed
    key-value metrics.

    Parameters
    ----------
    header_fill_color : str
        ARGB hex colour used for header row fills (default ``"FF1a3c6e"``).
    alt_row_color : str
        ARGB hex colour for alternating row shading (default ``"FFeaf0f9"``).
    """

    def __init__(
        self,
        header_fill_color: str = "FF1a3c6e",
        alt_row_color: str = "FFeaf0f9",
    ) -> None:
        self._hdr_fill = PatternFill(start_color=header_fill_color,
                                     end_color=header_fill_color,
                                     fill_type="solid")
        self._alt_fill = PatternFill(start_color=alt_row_color,
                                     end_color=alt_row_color,
                                     fill_type="solid")
        self._hdr_font = Font(name="Calibri", bold=True, color="FFFFFFFF",
                              size=11)
        self._body_font = Font(name="Calibri", size=10)
        self._cat_font = Font(name="Calibri", bold=True, size=12,
                              color="FF1a3c6e")
        self._thin_border = Border(
            left=Side(style="thin", color="FFCCCCCC"),
            right=Side(style="thin", color="FFCCCCCC"),
            top=Side(style="thin", color="FFCCCCCC"),
            bottom=Side(style="thin", color="FFCCCCCC"),
        )
        self._center = Alignment(horizontal="center", vertical="center")
        self._wrap = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate_workbook(
        self,
        results: List[Dict],
        output_path: str,
    ) -> str:
        """Generate an Excel workbook from analysis results.

        Creates:

        1. A **Summary** sheet with columns: ``#``, ``Category``,
           ``Method``, ``Timestamp``, and the first several key metrics.
        2. One sheet per *category* containing detailed flattened
           results for every analysis in that category.

        Parameters
        ----------
        results : list[dict]
            Analysis history records (``category``, ``method``,
            ``result``, ``timestamp``).
        output_path : str
            Destination path for the ``.xlsx`` file.

        Returns
        -------
        str
            The *output_path* on success.

        Raises
        ------
        ValueError
            If *results* is empty.
        """
        if not results:
            raise ValueError("No analysis results to export.")
        _ensure_parent_dir(output_path)

        wb = Workbook()
        self._write_summary_sheet(wb, results)
        self._write_category_sheets(wb, results)
        wb.save(output_path)
        return output_path

    def generate_data_sheet(
        self,
        data: Dict[str, np.ndarray],
        output_path: str,
        sheet_name: str = "Data",
    ) -> str:
        """Export raw data arrays to an Excel sheet.

        Each key in *data* becomes a column header and the corresponding
        NumPy array (1-D) provides the column values.

        Parameters
        ----------
        data : dict[str, numpy.ndarray]
            Mapping of column names to 1-D arrays.
        output_path : str
            Destination path for the ``.xlsx`` file.
        sheet_name : str
            Name of the worksheet (default ``"Data"``).

        Returns
        -------
        str
            The *output_path* on success.
        """
        if not data:
            raise ValueError("No data arrays to export.")
        _ensure_parent_dir(output_path)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit

        # Header row
        col_idx = 1
        headers = list(data.keys())
        for hdr in headers:
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.font = self._hdr_font
            cell.fill = self._hdr_fill
            cell.alignment = self._center
            cell.border = self._thin_border
            col_idx += 1

        # Determine max length across all arrays
        max_len = max(len(v) for v in data.values() if hasattr(v, "__len__"))

        # Data rows
        for row_idx in range(2, max_len + 2):
            for col_offset, key in enumerate(headers):
                arr = data[key]
                if hasattr(arr, "__len__") and row_idx - 2 < len(arr):
                    val = _excel_val(arr[row_idx - 2])
                else:
                    val = None
                cell = ws.cell(row=row_idx, column=col_offset + 1, value=val)
                cell.font = self._body_font
                cell.alignment = self._wrap
                cell.border = self._thin_border
                if isinstance(val, float):
                    cell.number_format = "0.0000"
                if (row_idx - 2) % 2 == 1:
                    cell.fill = self._alt_fill

        # Auto-fit column widths
        for col_offset, key in enumerate(headers):
            max_w = max(len(str(key)), 12)
            arr = data[key]
            if hasattr(arr, "__len__"):
                for j in range(min(len(arr), 20)):
                    w = len(_fmt_val(arr[j]))
                    if w > max_w:
                        max_w = w
            ws.column_dimensions[get_column_letter(col_offset + 1)].width = min(
                max_w + 2, 40
            )

        wb.save(output_path)
        return output_path

    # ------------------------------------------------------------------
    # Internal: Excel building blocks
    # ------------------------------------------------------------------

    def _write_summary_sheet(
        self, wb: Workbook, results: List[Dict]
    ) -> None:
        """Create the Summary sheet with one row per analysis."""
        ws = wb.active
        ws.title = "Summary"

        # Collect all unique metric keys across results for extra columns
        all_keys: List[str] = []
        seen: set = set()
        for rec in results:
            res = rec.get("result", {})
            if isinstance(res, dict):
                flat = _flatten_dict(res)
                for k in flat:
                    if k not in seen:
                        all_keys.append(k)
                        seen.add(k)

        # Limit extra metric columns to keep summary readable
        extra_keys = all_keys[:10]

        headers = ["#", "Category", "Method", "Timestamp"] + extra_keys
        for ci, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = self._hdr_font
            cell.fill = self._hdr_fill
            cell.alignment = self._center
            cell.border = self._thin_border

        for ri, rec in enumerate(results, 2):
            row_num = ri - 1
            res = rec.get("result", {})
            flat = _flatten_dict(res) if isinstance(res, dict) else {}
            values = [
                row_num,
                rec.get("category", ""),
                rec.get("method", ""),
                rec.get("timestamp", "")[:19],
            ]
            for k in extra_keys:
                values.append(_excel_val(flat.get(k)))
            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = self._body_font
                cell.alignment = self._wrap
                cell.border = self._thin_border
                if isinstance(val, float):
                    cell.number_format = "0.0000"
                if row_num % 2 == 0:
                    cell.fill = self._alt_fill

        # Column widths
        widths = [5, 22, 26, 20] + [18] * len(extra_keys)
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    def _write_category_sheets(
        self, wb: Workbook, results: List[Dict]
    ) -> None:
        """Create one sheet per category with detailed results."""
        groups = _group_by_category(results)
        for cat, recs in groups.items():
            sheet_name = cat[:31]  # Excel limit
            ws = wb.create_sheet(title=sheet_name)

            # Collect all keys for this category
            all_keys: List[str] = []
            seen: set = set()
            for rec in recs:
                res = rec.get("result", {})
                flat = _flatten_dict(res) if isinstance(res, dict) else {}
                for k in flat:
                    if k not in seen:
                        all_keys.append(k)
                        seen.add(k)

            headers = ["Method", "Timestamp"] + all_keys
            for ci, hdr in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=hdr)
                cell.font = self._hdr_font
                cell.fill = self._hdr_fill
                cell.alignment = self._center
                cell.border = self._thin_border

            for ri, rec in enumerate(recs, 2):
                res = rec.get("result", {})
                flat = _flatten_dict(res) if isinstance(res, dict) else {}
                values = [
                    rec.get("method", ""),
                    rec.get("timestamp", "")[:19],
                ]
                for k in all_keys:
                    values.append(_excel_val(flat.get(k)))
                for ci, val in enumerate(values, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = self._body_font
                    cell.alignment = self._wrap
                    cell.border = self._thin_border
                    if isinstance(val, float):
                        cell.number_format = "0.0000"
                    if (ri - 2) % 2 == 1:
                        cell.fill = self._alt_fill

            # Column widths
            col_widths = [28, 20] + [max(18, min(len(k) + 4, 40))
                                     for k in all_keys]
            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

            # Freeze top row
            ws.freeze_panes = "A2"

        # Freeze top row on summary as well
        if "Summary" in wb.sheetnames:
            wb["Summary"].freeze_panes = "A2"
