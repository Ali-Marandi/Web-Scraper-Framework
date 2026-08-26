"""
WebScraper Pro - Data Exporter
Exports scraped data to CSV, JSON, Excel, XML, and SQL formats.
"""

import csv
import json
import os
import sqlite3
import xml.etree.ElementTree as ET
from xml.dom import minidom
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path
from io import StringIO


class DataExporter:
    """
    Multi-format data exporter for scraped results.
    
    Supported formats:
    - CSV (with BOM for Excel compatibility)
    - JSON (pretty-printed)
    - Excel (XLSX with formatting)
    - XML (structured)
    - SQL (SQLite database)
    - HTML table
    """

    FORMAT_CSV = "csv"
    FORMAT_JSON = "json"
    FORMAT_EXCEL = "xlsx"
    FORMAT_XML = "xml"
    FORMAT_SQL = "sql"
    FORMAT_HTML = "html"

    def __init__(self):
        self._last_export_path: Optional[str] = None
        self._export_count = 0

    @property
    def last_export_path(self) -> Optional[str]:
        return self._last_export_path

    @property
    def export_count(self) -> int:
        return self._export_count

    def export(self, data: List[Dict[str, Any]], format: str,
               filepath: str, **kwargs) -> str:
        """
        Export data to the specified format.
        Returns the path to the created file.
        """
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

        exporters = {
            self.FORMAT_CSV: self._export_csv,
            self.FORMAT_JSON: self._export_json,
            self.FORMAT_EXCEL: self._export_excel,
            self.FORMAT_XML: self._export_xml,
            self.FORMAT_SQL: self._export_sql,
            self.FORMAT_HTML: self._export_html,
        }

        exporter = exporters.get(format)
        if not exporter:
            raise ValueError(f"Unsupported format: {format}. Supported: {list(exporters.keys())}")

        result_path = exporter(data, filepath, **kwargs)
        self._last_export_path = result_path
        self._export_count += 1
        return result_path

    def _export_csv(self, data: List[Dict], filepath: str, **kwargs) -> str:
        """Export to CSV with UTF-8 BOM for Excel compatibility."""
        if not data:
            # Write empty CSV with headers if provided
            headers = kwargs.get("headers", [])
            with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
                if headers:
                    writer = csv.writer(f)
                    writer.writerow(headers)
            return filepath

        # Collect all field names preserving order
        fieldnames = []
        for row in data:
            for key in row.keys():
                if key not in fieldnames and not key.startswith("_"):
                    fieldnames.append(key)

        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for row in data:
                cleaned = {}
                for k, v in row.items():
                    if not k.startswith("_"):
                        if isinstance(v, list):
                            cleaned[k] = "; ".join(str(i) for i in v)
                        else:
                            cleaned[k] = v
                writer.writerow(cleaned)

        return filepath

    def _export_json(self, data: List[Dict], filepath: str, **kwargs) -> str:
        """Export to JSON with pretty printing."""
        # Remove internal fields
        clean_data = []
        for row in data:
            cleaned = {k: v for k, v in row.items() if not k.startswith("_")}
            clean_data.append(cleaned)

        indent = kwargs.get("indent", 2)
        ensure_ascii = kwargs.get("ensure_ascii", False)

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(clean_data, f, indent=indent, ensure_ascii=ensure_ascii, default=str)

        return filepath

    def _export_excel(self, data: List[Dict], filepath: str, **kwargs) -> str:
        """Export to Excel XLSX with formatting."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return self._export_csv(data, filepath.replace(".xlsx", ".csv"), **kwargs)

        wb = Workbook()
        ws = wb.active
        ws.title = kwargs.get("sheet_name", "Scraped Data")

        if not data:
            headers = kwargs.get("headers", ["No Data"])
            ws.append(headers)
            wb.save(filepath)
            return filepath

        # Collect headers
        headers = []
        for row in data:
            for key in row.keys():
                if key not in headers and not key.startswith("_"):
                    headers.append(key)

        # Styles
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        cell_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row.get(header, "")
                if isinstance(value, list):
                    value = "; ".join(str(v) for v in value)
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in data[:100]:  # Sample first 100 rows
                val = str(row.get(header, ""))
                max_length = max(max_length, min(len(val), 50))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 4

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add auto-filter
        if headers:
            ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(data) + 1}"

        wb.save(filepath)
        return filepath

    def _export_xml(self, data: List[Dict], filepath: str, **kwargs) -> str:
        """Export to structured XML."""
        root_tag = kwargs.get("root_tag", "scraped_data")
        item_tag = kwargs.get("item_tag", "item")
        
        root = ET.Element(root_tag)
        root.set("exported", datetime.now().isoformat())
        root.set("count", str(len(data)))

        for row in data:
            item = ET.SubElement(root, item_tag)
            for key, value in row.items():
                if not key.startswith("_"):
                    # Sanitize key for XML tag name
                    tag_name = re.sub(r"[^a-zA-Z0-9_]", "_", key) if key else "field"
                    if tag_name[0].isdigit():
                        tag_name = "f_" + tag_name
                    child = ET.SubElement(item, tag_name)
                    if isinstance(value, list):
                        for v in value:
                            sub = ET.SubElement(child, "value")
                            sub.text = str(v)
                    else:
                        child.text = str(value) if value else ""

        # Pretty print
        xml_str = ET.tostring(root, encoding="unicode")
        dom = minidom.parseString(xml_str)
        pretty = dom.toprettyxml(indent="  ")

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(pretty)

        return filepath

    def _export_sql(self, data: List[Dict], filepath: str, **kwargs) -> str:
        """Export to SQLite database."""
        table_name = kwargs.get("table_name", "scraped_data").replace(" ", "_")
        
        conn = sqlite3.connect(filepath)
        cursor = conn.cursor()

        if not data:
            conn.close()
            return filepath

        # Collect columns
        columns = []
        for row in data:
            for key in row.keys():
                if key not in columns and not key.startswith("_"):
                    columns.append(key)

        # Create table
        col_defs = ", ".join(f'"{c}" TEXT' for c in columns)
        cursor.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({col_defs})')

        # Insert data
        placeholders = ", ".join("?" for _ in columns)
        for row in data:
            values = []
            for col in columns:
                val = row.get(col, "")
                if isinstance(val, list):
                    val = json.dumps(val)
                values.append(str(val) if val else "")
            cursor.execute(f'INSERT INTO "{table_name}" VALUES ({placeholders})', values)

        # Add metadata table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS "_metadata" (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')
        cursor.execute('INSERT OR REPLACE INTO "_metadata" VALUES (?, ?)',
                       ("exported_at", datetime.now().isoformat()))
        cursor.execute('INSERT OR REPLACE INTO "_metadata" VALUES (?, ?)',
                       ("row_count", str(len(data))))
        cursor.execute('INSERT OR REPLACE INTO "_metadata" VALUES (?, ?)',
                       ("columns", json.dumps(columns)))

        conn.commit()
        conn.close()
        return filepath

    def _export_html(self, data: List[Dict], filepath: str, **kwargs) -> str:
        """Export to HTML table."""
        title = kwargs.get("title", "Scraped Data")
        
        if not data:
            html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>{title}</title>
<style>body{{font-family:Segoe UI,sans-serif;padding:20px;}}</style>
</head><body><h1>{title}</h1><p>No data to display.</p></body></html>"""
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(html)
            return filepath

        headers = []
        for row in data:
            for key in row.keys():
                if key not in headers and not key.startswith("_"):
                    headers.append(key)

        rows_html = ""
        for row in data:
            cells = ""
            for h in headers:
                val = row.get(h, "")
                if isinstance(val, list):
                    val = "; ".join(str(v) for v in val)
                cells += f"<td>{str(val)}</td>"
            rows_html += f"<tr>{cells}</tr>"

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{title}</title>
<style>
body {{ font-family: 'Segoe UI', Tahoma, sans-serif; padding: 20px; background: #f5f5f5; }}
h1 {{ color: #2F5496; }}
table {{ border-collapse: collapse; width: 100%; background: white; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
th {{ background: #2F5496; color: white; padding: 12px 16px; text-align: left; font-size: 14px; }}
td {{ padding: 10px 16px; border-bottom: 1px solid #e0e0e0; font-size: 13px; }}
tr:nth-child(even) {{ background: #f8f9fa; }}
tr:hover {{ background: #e8f0fe; }}
.info {{ color: #666; margin-bottom: 16px; font-size: 14px; }}
</style>
</head>
<body>
<h1>{title}</h1>
<p class="info">Exported on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | {len(data)} records</p>
<table>
<thead><tr>{''.join(f'<th>{h}</th>' for h in headers)}</tr></thead>
<tbody>{rows_html}</tbody>
</table>
</body></html>"""

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)
        return filepath

    def get_preview(self, data: List[Dict], max_rows: int = 10) -> str:
        """Generate a text preview of the data."""
        if not data:
            return "No data to preview."

        headers = []
        for row in data:
            for key in row.keys():
                if key not in headers and not key.startswith("_"):
                    headers.append(key)

        # Calculate column widths
        widths = {h: len(h) for h in headers}
        for row in data[:max_rows]:
            for h in headers:
                val = str(row.get(h, ""))
                if isinstance(row.get(h), list):
                    val = "; ".join(str(v) for v in row.get(h, []))
                widths[h] = max(widths[h], min(len(val), 40))

        # Build table
        lines = []
        header_line = " | ".join(h.ljust(widths[h]) for h in headers)
        separator = "-+-".join("-" * widths[h] for h in headers)
        lines.append(header_line)
        lines.append(separator)

        for row in data[:max_rows]:
            cells = []
            for h in headers:
                val = row.get(h, "")
                if isinstance(val, list):
                    val = "; ".join(str(v) for v in val)
                cells.append(str(val)[:widths[h]].ljust(widths[h]))
            lines.append(" | ".join(cells))

        if len(data) > max_rows:
            lines.append(f"\n... and {len(data) - max_rows} more rows")

        return "\n".join(lines)
